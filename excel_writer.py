from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from cas_parser import to_number

HEADER_FILL = PatternFill(start_color='8B1E2D', end_color='8B1E2D', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(bold=True, size=13)


def _write_table(ws, headers, rows, number_cols=(), start_row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r, row in enumerate(rows, start=start_row + 1):
        for c, h in enumerate(headers, start=1):
            val = row.get(h, '')
            if h in number_cols:
                num = to_number(val) if isinstance(val, str) else val
                ws.cell(row=r, column=c, value=num)
            else:
                ws.cell(row=r, column=c, value=val)
    for c, h in enumerate(headers, start=1):
        width = max(12, min(45, len(h) + 4))
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def write_workbook(parsed, out_path, source_filename=''):
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = 'NSDL CAS Summary'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = 'Source file'
    ws['B2'] = source_filename
    ws['A3'] = 'Consolidated Portfolio Value'
    ws['B3'] = parsed['summary'].get('consolidated_portfolio_value')
    ws['B3'].number_format = '#,##0.00'
    ws['A5'] = 'Asset Class'
    ws['B5'] = 'Value'
    ws['A5'].font = HEADER_FONT
    ws['B5'].font = HEADER_FONT
    ws['A5'].fill = HEADER_FILL
    ws['B5'].fill = HEADER_FILL
    row = 6
    for cls, val in parsed['summary'].get('portfolio_composition', {}).items():
        ws.cell(row=row, column=1, value=cls)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = '#,##0.00'
        row += 1
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 20

    # --- Holdings sheets, one per asset class ---
    holdings = parsed['holdings']
    by_class = {}
    for h in holdings:
        by_class.setdefault(h['asset_class'], []).append(h)

    class_titles = {
        'equities': 'Equities',
        'preference': 'Preference Shares',
        'mf_holdings': 'Mutual Funds (Demat)',
        'sgb': 'Sovereign Gold Bonds',
        'aif': 'Alternate Investment Fund',
    }
    class_headers = {
        'equities': ['account', 'holder', 'ISIN', 'Company Name', 'Face Value', 'No. of Shares',
                     'Market Price', 'Value'],
        'preference': ['account', 'holder', 'ISIN', 'Company Name', 'Face Value', 'No. of Shares',
                       'Market Price/Unit', 'Value'],
        'mf_holdings': ['account', 'holder', 'ISIN', 'ISIN Description', 'No. of Units', 'NAV', 'Value'],
        'sgb': ['account', 'holder', 'ISIN', 'Issuer Name', 'Coupon Rate/Frequency', 'Maturity Date',
                'No. of Units', 'Face Value/Unit', 'Market Price/unit', 'Value'],
        'aif': ['account', 'holder', 'ISIN', 'ISIN Description', 'No. of Units', 'NAV', 'Value'],
    }
    number_cols = {'Face Value', 'No. of Shares', 'Market Price', 'Value', 'Market Price/Unit',
                    'No. of Units', 'NAV', 'Face Value/Unit', 'Market Price/unit'}

    for cls in ['equities', 'preference', 'mf_holdings', 'sgb', 'aif']:
        rows = by_class.get(cls, [])
        if not rows:
            continue
        ws = wb.create_sheet(class_titles[cls][:31])
        _write_table(ws, class_headers[cls], rows, number_cols=number_cols)

    # --- Mutual Fund Folios sheet ---
    folios = parsed['mf_folios']
    if folios:
        ws = wb.create_sheet('Mutual Fund Folios')
        headers = ['ISIN/UCC', 'ISIN Description', 'Folio No.', 'No. of Units', 'Avg Cost/Unit',
                   'Total Cost', 'Current NAV/unit', 'Current Value', 'Unrealised P/L',
                   'Annualised Return %']
        num_cols = {'Folio No.', 'No. of Units', 'Avg Cost/Unit', 'Total Cost', 'Current NAV/unit',
                    'Current Value', 'Unrealised P/L', 'Annualised Return %'}
        _write_table(ws, headers, folios, number_cols=num_cols)

    wb.save(out_path)
    return out_path


def write_workbook_full(parsed, demat_txns, mf_txns, out_path, source_filename=''):
    """Same as write_workbook, plus two Transactions sheets."""
    write_workbook(parsed, out_path, source_filename)
    from openpyxl import load_workbook
    wb = load_workbook(out_path)

    if demat_txns:
        ws = wb.create_sheet('Demat Transactions')
        headers = ['account', 'holder', 'ISIN', 'Company/Scheme', 'Date', 'Order No', 'Description',
                   'Instruction Details', 'Opening Balance', 'Debit', 'Credit', 'Closing Balance']
        num_cols = {'Opening Balance', 'Debit', 'Credit', 'Closing Balance'}
        _write_table(ws, headers, demat_txns, number_cols=num_cols)

    if mf_txns:
        ws = wb.create_sheet('MF Transactions')
        headers = ['ISIN', 'Scheme', 'Folio No.', 'Date', 'Transaction Details', 'Amount',
                   'Stamp Duty', 'NAV', 'Price', 'Units']
        num_cols = {'Amount', 'Stamp Duty', 'NAV', 'Price', 'Units'}
        _write_table(ws, headers, mf_txns, number_cols=num_cols)

    wb.save(out_path)
    return out_path
